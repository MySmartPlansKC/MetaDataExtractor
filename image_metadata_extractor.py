# MetadataExtractor/image_metadata_extractor.py

import json
import logging
import openpyxl
import os
import pillow_heif
import re
import shutil
import subprocess
import sys
import time
from datetime import datetime
from exifread import process_file
from fractions import Fraction as Ratio
from PIL import Image, ImageDraw, ImageFont, ExifTags

# Increase the pixel limit to 300 million pixels
Image.MAX_IMAGE_PIXELS = 300000000

# Versioning
__version__ = "2.3.0"
# pyinstaller --onefile --icon=metadata.ico --name MetaData-V2.3.0 image_metadata_extractor.py
# pyinstaller --onefile --icon=metadata.ico --name MetaData-V2.3.0 --add-data "exiftool-13.09_64/exiftool.exe;exiftool-13.09_64" --add-data "exiftool-13.09_64/exiftool_files;exiftool-13.09_64/exiftool_files" --paths "exiftool-13.09_64/exiftool_files" image_metadata_extractor.py

script_dir = os.path.dirname(os.path.realpath(__file__))


def setup_logging():
    log_file = "metadata_extraction.log"

    if os.path.exists(log_file):
        response = input(
            f"Log file '{log_file}' already exists. Delete it? (y/n, default y): "
        )
        if response.lower() != "n":
            os.remove(log_file)
            print(f"Log file '{log_file}' deleted.")
        else:
            print("Log file will be appended.")

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler(),
        ],
    )


# Global Configuration
METADATA_HEADER = f"MySmartPlans MetaData Tracker v{__version__}\n\n"
EXIFTOOL_PATH = "./exiftool-13.09_64/exiftool"

# File tracking
METADATA_FORMAT = "xlsx"  # Choose (txt or xlsx)
CREATE_METADATA_FILE = True
WRITE_RAW_METADATA = False

# Dynamic Padding Configuration
PADDING_LEFT_FACTOR = 0.03
PADDING_RIGHT_FACTOR = 0.50
PADDING_TOP_FACTOR = 0.02
PADDING_BOTTOM_FACTOR = 0.01

NO_METADATA_MESSAGE = "No Metadata Available\n"

OVERLAY_POSITIONS = {1: "top-left", 2: "top-right", 3: "bottom-left", 4: "bottom-right"}


def get_base_path():
    if getattr(sys, "frozen", False):
        # If the application is frozen using PyInstaller
        return os.path.dirname(sys.executable)
    else:
        # Normal execution (e.g., script or interactive)
        return os.path.dirname(os.path.abspath(__file__))


def parse_image_date(date_str):
    try:
        return datetime.strptime(date_str, "%Y:%m:%d %H:%M:%S").strftime(
            "%Y-%m-%d %H:%M:%S"
        )
    except ValueError:
        return date_str


def format_filesize_kb(filesize_bytes):
    if filesize_bytes is None:
        return "N/A"

    filesize_kb = filesize_bytes / 1024
    return f"{filesize_kb:.2f} KB"


def convert_gps_to_dms(coordinates, reference):
    if coordinates is None or isinstance(coordinates, str):
        return coordinates or "GPS Data not found"

    def calculate_dms(data):
        if isinstance(data, list):
            return sum(float(x) for x in data)
        elif isinstance(data, Ratio):  # Check for Ratio objects
            return float(data)
        else:
            raise ValueError("Unexpected data type in GPS coordinates")

    degrees = calculate_dms(coordinates.values[0] if coordinates else None)
    minutes = calculate_dms(coordinates.values[1] if coordinates else None) / 60
    seconds = calculate_dms(coordinates.values[2] if coordinates else None) / 3600

    return f"{degrees:.1f}° {minutes:.3f}' {seconds:.4f}\" {reference}"


def process_heic(image_path, text, output_directory, overlay_position):
    try:
        # Register the HEIF opener with Pillow
        pillow_heif.register_heif_opener()

        # Pass the image to overlay_text for further processing and saving
        overlay_text(
            image_path=image_path,
            text=text,
            position=(10, 10),
            output_directory=output_directory,
            overlay_position=overlay_position,
        )

        logging.info(f"Successfully processed HEIC file: {image_path}")

    except Exception as e:
        logging.error(f"Error processing HEIC file: {image_path} - Error: {e}")


def get_image_metadata(image_path):
    try:
        # Construct the ExifTool command with the updated path
        # command = [EXIFTOOL_PATH, "-j", image_path]  # -j for JSON output

        exiftool_path = os.path.join(sys._MEIPASS, "exiftool-13.09_64", "exiftool.exe")
        command = [exiftool_path, "-j", image_path]

        # Execute the command and capture the output
        process = subprocess.run(command, capture_output=True, text=True)

        # Check if there's any output from ExifTool
        if not process.stdout:
            raise ValueError("No output from ExifTool")

        # Parse the JSON output
        metadata = json.loads(process.stdout)[0]

        # Format the metadata (adjust as needed based on ExifTool's output)
        formatted_metadata = {
            "GPS Latitude": metadata.get("GPSLatitude", ""),
            "GPS Longitude": metadata.get("GPSLongitude", ""),
            "Origin Date": parse_image_date(metadata.get("DateTimeOriginal", "")),
            "Offset Time": metadata.get("OffsetTime", ""),
            "Orientation": metadata.get("Orientation", ""),
            "Make": metadata.get("Make", ""),
            "Model": metadata.get("Model", ""),
            "Image Width": metadata.get("ImageWidth", ""),
            "Image Height": metadata.get("ImageHeight", ""),
            "Megapixels": metadata.get("Megapixels", ""),  # Now directly available
        }

        return formatted_metadata, metadata  # Return both formatted and raw metadata

    except Exception as e:
        print(f"Error processing image: {image_path} - Error: {e}")
        return {}


def write_raw_metadata(raw_metadata, output_path):
    if WRITE_RAW_METADATA:
        with open(output_path, "w") as f:
            for key, value in raw_metadata.items():
                f.write(f"{key}: {value}\n")


def write_metadata(metadata, output_path):
    if METADATA_FORMAT == "txt":
        with open(output_path, "w") as f:
            f.write(METADATA_HEADER)
            f.write(f"Filename: {os.path.basename(output_path)}\n")
            f.write("\n")
            for file_info in metadata["files"]:
                f.write(f"Filename: {file_info['filename']}\n")
                f.write(f"File Size: {format_filesize_kb(file_info['file_size'])}\n")
                f.write(f"File Type: {file_info['file_type']}\n")
                f.write(f"File Path: {file_info['file_path']}\n")
                # Write metadata values if present
                f.write(f"Make: {file_info['metadata'].get('Make')}\n")
                f.write(f"Model: {file_info['metadata'].get('Model')}\n")
                lat_ref = file_info["metadata"].get("GPS GPSLatitudeRef") or "N"
                latitude = convert_gps_to_dms(
                    file_info["metadata"].get("GPS Latitude"), lat_ref
                )
                f.write(f"Latitude: {latitude}\n")
                lon_ref = file_info["metadata"].get("GPS GPSLongitudeRef") or "E"
                longitude = convert_gps_to_dms(
                    file_info["metadata"].get("GPS Longitude"), lon_ref
                )
                f.write(f"Longitude: {longitude}\n")
                f.write(f"Origin Date: {file_info['metadata'].get('Origin Date')}\n")
                f.write(f"Orientation: {file_info['metadata'].get('Orientation')}\n")
                f.write(f"Image Width: {file_info['metadata'].get('Image Width')}\n")
                f.write(f"Image Height: {file_info['metadata'].get('Image Height')}\n")
                f.write(f"Offset Time: {file_info['metadata'].get('Offset Time')}\n")
                f.write(f"Megapixels: {file_info['metadata'].get('Megapixels')}\n")
                f.write("\n")

    elif METADATA_FORMAT == "xlsx":
        try:
            # Attempt to load existing workbook
            wb = openpyxl.load_workbook(output_path)
            sheet = wb.active
        except FileNotFoundError:
            # Workbook doesn't exist, create headers
            wb = openpyxl.Workbook()
            sheet = wb.active
            row = 1
            sheet.cell(row, 1).value = "Filename"
            sheet.cell(row, 2).value = "File Path"
            sheet.cell(row, 3).value = "Origin Date"
            sheet.cell(row, 4).value = "Offset Time"
            sheet.cell(row, 5).value = "Orientation"
            sheet.cell(row, 6).value = "Make"
            sheet.cell(row, 7).value = "Model"
            sheet.cell(row, 8).value = "File Size"
            sheet.cell(row, 9).value = "File Type"
            sheet.cell(row, 10).value = "GPS Latitude"
            sheet.cell(row, 11).value = "GPS Longitude"
            sheet.cell(row, 12).value = "Image Width"
            sheet.cell(row, 13).value = "Image Height"
            sheet.cell(row, 14).value = "Megapixels"

        # Process file info from the metadata
        for file_info in metadata["files"]:
            row = sheet.max_row + 1

            sheet.cell(row, 1).value = file_info["filename"]
            sheet.cell(row, 2).value = file_info["file_path"]

            # Convert 'Origin Date'
            origin_date = file_info["metadata"].get("Origin Date")
            if origin_date:
                formatted_date = parse_image_date(origin_date)
                sheet.cell(row, 3).value = formatted_date

            # Convert 'Offset Time'
            offset_time = file_info["metadata"].get("Offset Time")
            if isinstance(offset_time, str):
                sheet.cell(row, 4).value = offset_time
            else:
                sheet.cell(row, 4).value = str(offset_time)
            sheet.cell(row, 5).value = str(file_info["metadata"].get("Orientation"))

            # Handle potential None or invalid characters
            make = file_info["metadata"].get("Make")
            if make:
                make = make.encode("ascii", errors="ignore").decode()
                sheet.cell(row, 6).value = make

            # Handle potential None or invalid characters
            model = file_info["metadata"].get("Model")
            if model:
                model = model.encode("ascii", errors="ignore").decode()
                sheet.cell(row, 7).value = model
            sheet.cell(row, 8).value = format_filesize_kb(file_info["file_size"])
            sheet.cell(row, 9).value = file_info["file_type"]
            sheet.cell(row, 10).value = str(file_info["metadata"].get("GPS Latitude"))
            sheet.cell(row, 11).value = str(file_info["metadata"].get("GPS Longitude"))
            image_width = file_info["metadata"].get("Image Width")
            if image_width:
                # sheet.cell(row, 12).value = int(image_width.printable)
                sheet.cell(row, 12).value = int(str(image_width))
            image_height = file_info["metadata"].get("Image Height")
            if image_height:
                # sheet.cell(row, 13).value = int(image_height.printable)
                sheet.cell(row, 13).value = int(str(image_height))
            megapixels = file_info["metadata"].get("Megapixels")
            if megapixels:
                # sheet.cell(row, 14).value = float(megapixels.printable)
                sheet.cell(row, 14).value = float(str(megapixels))

        wb.save(output_path)
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            column_letter = column_cells[0].column_letter
            sheet.column_dimensions[column_letter].width = length
        wb.save(output_path)

    else:
        raise ValueError("Unsupported format")


def open_image_without_orientation(image_path):
    try:
        img = Image.open(image_path)

        # Skip orientation check for HEIC files
        if not image_path.lower().endswith(".heic"):
            exif = img.getexif()
            orientation = 1
            for orientation in ExifTags.TAGS.keys():
                if ExifTags.TAGS[orientation] == "Orientation":
                    break

            if exif[orientation] == 2:  # Horizontal flip
                img = img.transpose(Image.FLIP_LEFT_RIGHT)
            elif exif[orientation] == 3:  # 180 degree rotation
                img = img.transpose(Image.ROTATE_180)
            elif exif[orientation] == 4:  # Vertical flip
                img = img.transpose(Image.FLIP_TOP_BOTTOM)
            elif exif[orientation] == 5:  # Transpose (90 degree rotation + flip)
                img = img.transpose(Image.TRANSPOSE)
            elif exif[orientation] == 6:  # 90 degree rotation
                img = img.transpose(Image.ROTATE_270)
            elif exif[orientation] == 7:  # Transpose (270 degree rotation + flip)
                img = img.transpose(Image.TRANSVERSE)
            elif exif[orientation] == 8:  # 270 degree rotation
                img = img.transpose(Image.ROTATE_90)

        return img

    except (KeyError, AttributeError, OSError):
        # exif data missing or image corrupt
        print(f"Image file might be corrupt or missing EXIF data: {image_path}")
        return Image.open(image_path)


def overlay_text(
    image_path,
    text,
    position,
    output_directory,
    overlay_position="top-left",
    color=(0, 0, 0),
    background_color=(255, 255, 255, 128),
):
    img = open_image_without_orientation(image_path).convert("RGBA")

    # Create a new image for semi-transparent overlay
    overlay = Image.new("RGBA", img.size, (255, 255, 255, 0))
    overlay_draw = ImageDraw.Draw(overlay)

    width, height = img.size

    # Dynamic Font Scaling
    font_size = max(
        12, int(height * 0.02)
    )  # Base size of 12, scales with 2% of image height
    font = ImageFont.truetype("arial.ttf", font_size)

    # Dynamic Padding
    padding_left = int(min(width, height) * PADDING_LEFT_FACTOR)
    padding_right = int(padding_left * PADDING_RIGHT_FACTOR)
    padding_top = int(height * PADDING_TOP_FACTOR)
    padding_bottom = int(height * PADDING_BOTTOM_FACTOR)

    if not text:
        text = NO_METADATA_MESSAGE

    text = METADATA_HEADER + text

    text_width, text_height = overlay_draw.textbbox((0, 0), text, font=font)[2:]

    if overlay_position == "top-left":
        text_position = (position[0], position[1])
    elif overlay_position == "top-right":
        text_position = (width - text_width - padding_right, position[1])
    elif overlay_position == "bottom-left":
        text_position = (position[0], height - text_height - padding_bottom)
    elif overlay_position == "bottom-right":
        text_position = (
            width - text_width - padding_right,
            height - text_height - padding_bottom,
        )

    overlay_draw.rectangle(
        (
            (text_position[0] - padding_left, text_position[1] - padding_top),
            (
                text_position[0] + text_width + padding_right,
                text_position[1] + text_height + padding_bottom,
            ),
        ),
        fill=background_color,
    )
    overlay_draw.text((text_position[0], text_position[1]), text, fill=color, font=font)

    cropped_overlay = overlay.crop(
        (
            text_position[0] - padding_left,
            text_position[1] - padding_top,
            text_position[0] + text_width + padding_right,
            text_position[1] + text_height + padding_bottom,
        )
    )

    # Paste only the cropped area onto the original image
    img.paste(
        cropped_overlay,
        (text_position[0] - padding_left, text_position[1] - padding_top),
        mask=cropped_overlay,
    )

    date_match = re.search(r"Date/Time: (\d{4}-\d{2}-\d{2})", text)
    if date_match:
        date_str = date_match.group(1)  # Extract the date string
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")  # Convert to datetime object
        date_folder = date_obj.strftime("%m-%d-%Y")
    else:
        date_folder = datetime.now().strftime(
            "%m-%d-%Y"
        )  # Fallback to current date if not found

    date_folder_path = os.path.join(output_directory, date_folder)
    os.makedirs(date_folder_path, exist_ok=True)

    # Save to output directory
    filename = os.path.basename(image_path)
    new_name = os.path.splitext(filename)[0] + "_MD.png"
    output_path = os.path.join(date_folder_path, new_name)
    img.save(output_path)
    # time.sleep(2)
    os.remove(image_path)


def check_and_clear_directory(directory):
    if os.path.exists(directory):
        if os.listdir(directory):  # Check if the directory is not empty
            response = input(
                f"Directory {directory} is not empty. Delete all contents? (y/n, default y): "
            )
            if response.lower() != "n":
                # Clear the directory
                for filename in os.listdir(directory):
                    file_path = os.path.join(directory, filename)
                    try:
                        if os.path.isfile(file_path) or os.path.islink(file_path):
                            os.unlink(file_path)
                        elif os.path.isdir(file_path):
                            shutil.rmtree(file_path)
                        logging.info(f"All contents of {directory} have been deleted.")
                    except PermissionError:
                        logging.warning(
                            f"Could not delete file {file_path} because it is being used by another process."
                        )
                        print(
                            f"\nError: Could not delete {file_path}. Please ensure the file is closed and not being used by any other program.\n"
                        )
                        input("Press Enter to continue...")
            else:
                logging.error("Operation aborted by the user.")
                return False
    else:
        os.makedirs(directory, exist_ok=True)
    return True


def main():
    setup_logging()

    base_path = get_base_path()

    # input_directory = r"E:\Python\xPDFTestFiles\IMAGES_IN"
    # output_directory = r"E:\Python\xPDFTestFiles\IMAGES_OUT"
    # error_directory = r"E:\Python\xPDFTestFiles\IMAGES_ERROR"

    input_directory = os.path.join(base_path, "IMAGES_IN")
    output_directory = os.path.join(base_path, "IMAGES_OUT")
    error_directory = os.path.join(base_path, "IMAGES_ERROR")

    # User selects the overlay position
    print("Choose the orientation of the metadata overlay for this batch of files:")
    print("1: Top Left (Default)")
    print("2: Top Right")
    print("3: Bottom Left")
    print("4: Bottom Right")
    overlay_choice = input(
        "Enter your choice (1-4, or press Enter for default): "
    ).strip()
    if not overlay_choice:
        overlay_choice = "1"
    overlay_position = OVERLAY_POSITIONS.get(int(overlay_choice), "top-left")

    logging.info(f"Starting Metadata Extractor version {__version__}")
    logging.info(f"Using overlay position {overlay_choice}")
    time.sleep(1)

    # Check and possibly clear the output directory
    if not check_and_clear_directory(output_directory):
        return
    time.sleep(1)

    # Check and possibly clear the error directory
    if not check_and_clear_directory(error_directory):
        return
    time.sleep(1)

    if not os.path.exists(input_directory):
        logging.error(f"Input directory does not exist: {input_directory}")
        return

    # Check if there are any Image files in the directory
    if not any(
        file.lower().endswith((".jpg", ".jpeg", ".png", ".heic"))
        for file in os.listdir(input_directory)
    ):
        logging.error("No Images found in the input directory.")
        return

    # Create directory if not exist
    os.makedirs(output_directory, exist_ok=True)
    os.makedirs(error_directory, exist_ok=True)

    has_errors = False

    directory_metadata = {"files": []}

    for root, dirs, files in os.walk(input_directory):
        for filename in files:
            filepath = os.path.join(root, filename)
            try:
                logging.info(f"Processing file: {filename}")

                # Exclude temporary JPEG files from processing
                if filepath.lower().endswith(
                    (".jpg", ".jpeg", ".png", ".heic")
                ) and not filepath.endswith(("_MD.png", "_MD.txt", "_temp.jpg")):
                    formatted_metadata, raw_metadata = get_image_metadata(filepath)

                    if WRITE_RAW_METADATA:
                        output_filename = (
                            os.path.splitext(filename)[0] + "_metadata.txt"
                        )
                        output_path = os.path.join(output_directory, output_filename)
                        write_raw_metadata(raw_metadata, output_path)

                    file_stats = os.stat(filepath)
                    file_size = file_stats.st_size
                    file_type = os.path.splitext(filename)[1]

                    directory_metadata["files"].append(
                        {
                            "filename": filename,
                            "metadata": formatted_metadata,
                            "file_size": file_size,
                            "file_type": file_type,
                            "file_path": filepath,
                        }
                    )

                    overlay_text_content = f"Filename: {os.path.basename(filepath)}\n"

                    if all(value is None for value in formatted_metadata.values()):
                        overlay_text_content += "No Metadata Available\n"
                    else:
                        for key, value in formatted_metadata.items():
                            if key == "GPS Latitude" and value:
                                lat_ref = (
                                    formatted_metadata.get("GPS GPSLatitudeRef") or "N"
                                )
                                latitude = convert_gps_to_dms(value, lat_ref)
                                overlay_text_content += f"Latitude: {latitude}\n"
                            elif key == "GPS Longitude" and value:
                                lon_ref = (
                                    formatted_metadata.get("GPS GPSLongitudeRef") or "E"
                                )
                                longitude = convert_gps_to_dms(value, lon_ref)
                                overlay_text_content += f"Longitude: {longitude}\n"
                            elif key == "Origin Date" and value:
                                overlay_text_content += f"Date/Time: {value}\n"
                            elif key == "Offset Time" and value:
                                overlay_text_content += f"Offset Time: {value}\n"
                            elif key == "Orientation" and value:
                                overlay_text_content += f"Orientation: {value}\n"
                            elif key == "Make" and value:
                                overlay_text_content += f"Make: {value}\n"
                            elif key == "Model" and value:
                                overlay_text_content += f"Model: {value}\n"
                            elif key == "Image Width" and value:
                                image_width = value
                            elif key == "Image Height" and value and image_width:
                                overlay_text_content += (
                                    f"Image Size: {image_width} x {value}\n"
                                )
                                image_width = None
                            elif key == "Megapixel" and value:
                                overlay_text_content += f"Megapixel: {value}\n"

                    logging.info(f"Processed file: {filename}")

                    # Use process_heic for HEIC files, overlay_text for others
                    if filepath.lower().endswith(".heic"):
                        process_heic(
                            filepath,
                            overlay_text_content,
                            output_directory,
                            overlay_position,
                        )
                    else:
                        overlay_text(
                            filepath,
                            overlay_text_content,
                            (10, 10),
                            output_directory,
                            overlay_position,
                        )
                        logging.info(f"Successfully processed Image file: {filepath}")

            except Exception as e:
                # Moves files to IMAGES_ERROR
                # Use copy to avoid deleting test files
                error_file_path = os.path.join(error_directory, filename)
                os.rename(filepath, error_file_path)
                # shutil.copy(filepath, error_file_path)

                logging.error(f"Failed to process {filename}. Error: {e}")
                print(f"An error occurred: {e}")
                print(f"Check the error output folder: {error_directory}")
                input("Press Enter to acknowledge and continue...")
                has_errors = True

        # Handle metadata file creation
        if CREATE_METADATA_FILE:
            working_directory_name = os.path.basename(input_directory)
            directory_metadata_path = os.path.join(
                output_directory, working_directory_name + "_MD." + METADATA_FORMAT
            )
            write_metadata(directory_metadata, directory_metadata_path)
    if has_errors:
        logging.error(
            f"Processing completed with errors. Please check: {error_directory}"
        )

    # Testing only
    input("Press Enter to close this window...")


if __name__ == "__main__":
    main()
