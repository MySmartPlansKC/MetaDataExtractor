# MetadataExtractor/get_image_metadata.py

import argparse
import os
from exifread import process_file
from datetime import datetime
import openpyxl
from PIL import Image, ImageDraw, ImageFont, ExifTags
from fractions import Fraction as Ratio

script_dir = os.path.dirname(os.path.realpath(__file__))

# Global Configuration
METADATA_HEADER = "MySmartPlans MetaData Tracker\n\n"
INPUT_DIRECTORY_GLOBAL = os.path.join(script_dir, r"..\testFiles")  # TESTING
OUTPUT_DIRECTORY_GLOBAL = os.path.join(script_dir, r"..\output")  # TESTING
# INPUT_DIRECTORY = os.path.join(script_dir, r"L:\Fresno\Procore Files\Photos\Processed\2024-0422\Unclassified")
# OUTPUT_DIRECTORY = os.path.join(script_dir, r"L:\Fresno\Procore Files\Photos\Processed\2024-0422\metadata")

# File tracking
METADATA_FORMAT = "xlsx"  # Choose (txt or xlsx)
CREATE_METADATA_FILE = True
WRITE_RAW_METADATA = False

# Dynamic Padding Configuration
PADDING_LEFT_FACTOR = 0.03
PADDING_RIGHT_FACTOR = 0.50
PADDING_TOP_FACTOR = 0.02
PADDING_BOTTOM_FACTOR = 0.01

NO_METADATA_MESSAGE = "No Metadata Available\n"


def parse_image_date(date_str):
    try:
        return datetime.strptime(date_str, "%Y:%m:%d %H:%M:%S").strftime(
            "%Y-%m-%d %H:%M:%S"
        )
    except ValueError:
        return date_str


def format_filesize_kb(filesize_bytes):
    if filesize_bytes is None:
        return "N/A"

    filesize_kb = filesize_bytes / 1024
    return f"{filesize_kb:.2f} KB"


def convert_gps_to_dms(coordinates, reference):
    if coordinates is None:
        return "GPS Data not found"

    def calculate_dms(data):
        if isinstance(data, list):
            return sum(float(x) for x in data)
        elif isinstance(data, Ratio):  # Check for Ratio objects
            return float(data)
        else:
            raise ValueError("Unexpected data type in GPS coordinates")

    degrees = calculate_dms(coordinates.values[0] if coordinates else None)
    minutes = calculate_dms(coordinates.values[1] if coordinates else None) / 60
    seconds = calculate_dms(coordinates.values[2] if coordinates else None) / 3600

    return f"{degrees:.0f}° {minutes:.0f}' {seconds:.2f}\" {reference}"


def get_image_metadata(image_path):
    try:
        with open(image_path, "rb") as file:
            tags = process_file(file)
        metadata = {
            "GPS Latitude": tags.get("GPS GPSLatitude") or None,
            "GPS Longitude": tags.get("GPS GPSLongitude") or None,
            "Origin Date": (
                parse_image_date(str(tags.get("EXIF DateTimeOriginal")))
                if tags.get("EXIF DateTimeOriginal")
                else None
            ),
            "Offset Time": (
                tags.get("EXIF OffsetTime") if tags.get("EXIF OffsetTime") else None
            ),
            "Orientation": tags.get("Image Orientation") or None,
            "Make": (
                str(tags.get("Image Make").printable).strip()
                if tags.get("Image Make")
                else None
            ),
            "Model": (
                str(tags.get("Image Model").printable).strip()
                if tags.get("Image Model")
                else None
            ),
            "Image Width": (
                tags.get("EXIF ExifImageWidth")
                if tags.get("EXIF ExifImageWidth")
                else None
            ),
            "Image Height": (
                tags.get("EXIF ExifImageLength")
                if tags.get("EXIF ExifImageLength")
                else None
            ),
            "Megapixels": tags.get("Megapixels") if tags.get("Megapixels") else None,
        }
        return metadata, tags
    except Exception as e:
        print(f"Error processing image: {image_path} - Error: {e}")
        return {}


def write_raw_metadata(raw_metadata, output_path):
    if WRITE_RAW_METADATA:
        with open(output_path, "w") as f:
            for key, value in raw_metadata.items():
                f.write(f"{key}: {value}\n")


def write_metadata(metadata, output_path):
    if METADATA_FORMAT == "txt":
        with open(output_path, "w") as f:
            f.write(METADATA_HEADER)
            f.write(f"Filename: {os.path.basename(output_path)}\n")
            f.write("\n")
            for file_info in metadata["files"]:
                f.write(f"Filename: {file_info['filename']}\n")
                f.write(f"File Size: {format_filesize_kb(file_info['file_size'])}\n")
                f.write(f"File Type: {file_info['file_type']}\n")
                f.write(f"File Path: {file_info['file_path']}\n")
                # Write metadata values if present
                f.write(f"Make: {file_info['metadata'].get('Make')}\n")
                f.write(f"Model: {file_info['metadata'].get('Model')}\n")
                lat_ref = file_info["metadata"].get("GPS GPSLatitudeRef") or "N"
                latitude = convert_gps_to_dms(
                    file_info["metadata"].get("GPS Latitude"), lat_ref
                )
                f.write(f"Latitude: {latitude}\n")
                lon_ref = file_info["metadata"].get("GPS GPSLongitudeRef") or "E"
                longitude = convert_gps_to_dms(
                    file_info["metadata"].get("GPS Longitude"), lon_ref
                )
                f.write(f"Longitude: {longitude}\n")
                f.write(f"Origin Date: {file_info['metadata'].get('Origin Date')}\n")
                f.write(f"Orientation: {file_info['metadata'].get('Orientation')}\n")
                f.write(f"Image Width: {file_info['metadata'].get('Image Width')}\n")
                f.write(f"Image Height: {file_info['metadata'].get('Image Height')}\n")
                f.write(f"Offset Time: {file_info['metadata'].get('Offset Time')}\n")
                f.write(f"Megapixels: {file_info['metadata'].get('Megapixels')}\n")
                f.write("\n")

    elif METADATA_FORMAT == "xlsx":
        try:
            # Attempt to load existing workbook
            wb = openpyxl.load_workbook(output_path)
            sheet = wb.active
        except FileNotFoundError:
            # Workbook doesn't exist, create headers
            wb = openpyxl.Workbook()
            sheet = wb.active
            row = 1
            sheet.cell(row, 1).value = "Filename"
            sheet.cell(row, 2).value = "File Path"
            sheet.cell(row, 3).value = "Origin Date"
            sheet.cell(row, 4).value = "Offset Time"
            sheet.cell(row, 5).value = "Orientation"
            sheet.cell(row, 6).value = "Make"
            sheet.cell(row, 7).value = "Model"
            sheet.cell(row, 8).value = "File Size"
            sheet.cell(row, 9).value = "File Type"
            sheet.cell(row, 10).value = "GPS Latitude"
            sheet.cell(row, 11).value = "GPS Longitude"
            sheet.cell(row, 12).value = "Image Width"
            sheet.cell(row, 13).value = "Image Height"
            sheet.cell(row, 14).value = "Megapixels"

        # Process file info from the metadata
        for file_info in metadata["files"]:
            row = sheet.max_row + 1

            sheet.cell(row, 1).value = file_info["filename"]
            sheet.cell(row, 2).value = file_info["file_path"]

            # Convert 'Origin Date'
            origin_date = file_info["metadata"].get("Origin Date")
            if origin_date:
                formatted_date = parse_image_date(origin_date)
                sheet.cell(row, 3).value = formatted_date

            # Convert 'Offset Time'
            offset_time = file_info["metadata"].get("Offset Time")
            if isinstance(offset_time, str):
                sheet.cell(row, 4).value = offset_time
            else:
                sheet.cell(row, 4).value = str(offset_time)
            sheet.cell(row, 5).value = str(file_info["metadata"].get("Orientation"))

            # Handle potential None or invalid characters
            make = file_info["metadata"].get("Make")
            if make:
                make = make.encode("ascii", errors="ignore").decode()
                sheet.cell(row, 6).value = make

            # Handle potential None or invalid characters
            model = file_info["metadata"].get("Model")
            if model:
                model = model.encode("ascii", errors="ignore").decode()
                sheet.cell(row, 7).value = model
            sheet.cell(row, 8).value = format_filesize_kb(file_info["file_size"])
            sheet.cell(row, 9).value = file_info["file_type"]
            sheet.cell(row, 10).value = str(file_info["metadata"].get("GPS Latitude"))
            sheet.cell(row, 11).value = str(file_info["metadata"].get("GPS Longitude"))
            image_width = file_info["metadata"].get("Image Width")
            if image_width:
                sheet.cell(row, 12).value = int(image_width.printable)
            image_height = file_info["metadata"].get("Image Height")
            if image_height:
                sheet.cell(row, 13).value = int(image_height.printable)
            megapixels = file_info["metadata"].get("Megapixels")
            if megapixels:
                sheet.cell(row, 14).value = float(megapixels.printable)

        wb.save(output_path)
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            column_letter = column_cells[0].column_letter
            sheet.column_dimensions[column_letter].width = length
        wb.save(output_path)

    else:
        raise ValueError("Unsupported format")


def open_image_without_orientation(image_path):
    try:
        img = Image.open(image_path)
        exif = img.getexif()
        orientation = 1
        for orientation in ExifTags.TAGS.keys():
            if ExifTags.TAGS[orientation] == "Orientation":
                break

        if exif[orientation] == 2:  # Horizontal flip
            img = img.transpose(Image.FLIP_LEFT_RIGHT)
        elif exif[orientation] == 3:  # 180 degree rotation
            img = img.transpose(Image.ROTATE_180)
        elif exif[orientation] == 4:  # Vertical flip
            img = img.transpose(Image.FLIP_TOP_BOTTOM)
        elif exif[orientation] == 5:  # Transpose (90 degree rotation + flip)
            img = img.transpose(Image.TRANSPOSE)
        elif exif[orientation] == 6:  # 90 degree rotation
            img = img.transpose(Image.ROTATE_270)
        elif exif[orientation] == 7:  # Transpose (270 degree rotation + flip)
            img = img.transpose(Image.TRANSVERSE)
        elif exif[orientation] == 8:  # 270 degree rotation
            img = img.transpose(Image.ROTATE_90)

        return img
    except (KeyError, AttributeError, OSError):
        # exif data missing or image corrupt
        print(f"Image file might be corrupt or missing EXIF data: {image_path}")
        return Image.open(image_path)


def overlay_text(
    image_path,
    text,
    position,
    output_directory,
    color=(0, 0, 0),
    background_color=(255, 255, 255, 128),
):
    img = open_image_without_orientation(image_path).convert("RGBA")

    # Create a new image for semi-transparent overlay
    overlay = Image.new("RGBA", img.size, (255, 255, 255, 0))
    overlay_draw = ImageDraw.Draw(overlay)

    width, height = img.size

    # Dynamic Font Scaling
    font_size = max(
        12, int(height * 0.02)
    )  # Base size of 12, scales with 2% of image height
    font = ImageFont.truetype("arial.ttf", font_size)

    # Dynamic Padding
    padding_left = int(min(width, height) * PADDING_LEFT_FACTOR)
    padding_right = int(padding_left * PADDING_RIGHT_FACTOR)
    padding_top = int(height * PADDING_TOP_FACTOR)
    padding_bottom = int(height * PADDING_BOTTOM_FACTOR)

    if not text:
        text = NO_METADATA_MESSAGE

    text = METADATA_HEADER + text

    text_width, text_height = overlay_draw.textbbox((0, 0), text, font=font)[2:]
    overlay_draw.rectangle(
        (
            (position[0] - padding_left, position[1] - padding_top),
            (
                position[0] + text_width + padding_right,
                position[1] + text_height + padding_bottom,
            ),
        ),
        fill=background_color,
    )
    overlay_draw.text((position[0], position[1]), text, fill=color, font=font)

    cropped_overlay = overlay.crop(
        (
            position[0] - padding_left,
            position[1] - padding_top,
            position[0] + text_width + padding_right,
            position[1] + text_height + padding_bottom,
        )
    )

    # Paste only the cropped area onto the original image
    img.paste(
        cropped_overlay,
        (position[0] - padding_left, position[1] - padding_top),
        mask=cropped_overlay,
    )

    # Save to output directory
    filename = os.path.basename(image_path)
    new_name = os.path.splitext(filename)[0] + "_MD.png"
    output_path = os.path.join(output_directory, new_name)
    img.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Process images and extract metadata")
    parser.add_argument("-i", "--input_dir", required=False, help="Input directory")
    parser.add_argument("-o", "--output_dir", required=False, help="Output directory")
    args = parser.parse_args()

    # Prioritize command-line arguments if provided, otherwise use globals
    INPUT_DIRECTORY = args.input_dir or INPUT_DIRECTORY_GLOBAL
    OUTPUT_DIRECTORY = args.output_dir or OUTPUT_DIRECTORY_GLOBAL

    # Create directory if not exist
    os.makedirs(OUTPUT_DIRECTORY, exist_ok=True)

    directory_metadata = {"files": []}  # List for storing file-level metadata

    for root, dirs, files in os.walk(INPUT_DIRECTORY):
        for filename in files:
            filepath = os.path.join(root, filename)

            print(f"Processing file: {filename}")

            if filepath.lower().endswith(
                (".jpg", ".jpeg", ".png")
            ) and not filepath.endswith(("_MD.png", "_MD.txt")):
                formatted_metadata, raw_metadata = get_image_metadata(filepath)

                if WRITE_RAW_METADATA:  # Conditionally write raw data
                    output_filename = os.path.splitext(filename)[0] + "_metadata.txt"
                    output_path = os.path.join(OUTPUT_DIRECTORY, output_filename)
                    write_raw_metadata(raw_metadata, output_path)

                # Get file information
                file_stats = os.stat(filepath)
                file_size = file_stats.st_size
                file_type = os.path.splitext(filename)[1]  # Get extension

                # Add file info to the metadata
                directory_metadata["files"].append(
                    {
                        "filename": filename,
                        "metadata": formatted_metadata,
                        "file_size": file_size,
                        "file_type": file_type,
                        "file_path": filepath,
                    }
                )

                overlay_text_content = f"Filename: {os.path.basename(filepath)}\n"

                if all(value is None for value in formatted_metadata.values()):
                    overlay_text_content += "No Metadata Available\n"
                else:
                    for key, value in formatted_metadata.items():
                        if key == "GPS Latitude" and value:
                            lat_ref = (
                                formatted_metadata.get("GPS GPSLatitudeRef") or "N"
                            )
                            latitude = convert_gps_to_dms(value, lat_ref)
                            overlay_text_content += f"Latitude: {latitude}\n"
                        elif key == "GPS Longitude" and value:
                            lon_ref = (
                                formatted_metadata.get("GPS GPSLongitudeRef") or "E"
                            )
                            longitude = convert_gps_to_dms(value, lon_ref)
                            overlay_text_content += f"Longitude: {longitude}\n"
                        elif key == "Origin Date" and value:
                            overlay_text_content += f"Date/Time: {value}\n"
                        elif key == "Offset Time" and value:
                            overlay_text_content += f"Offset Time: {value}\n"
                        elif key == "Orientation" and value:
                            overlay_text_content += f"Orientation: {value}\n"
                        elif key == "Make" and value:
                            overlay_text_content += f"Make: {value}\n"
                        elif key == "Model" and value:
                            overlay_text_content += f"Model: {value}\n"
                        elif key == "Image Width" and value:
                            image_width = value
                        elif key == "Image Height" and value and image_width:
                            overlay_text_content += (
                                f"Image Size: {image_width} x {value}\n"
                            )
                            image_width = None
                        elif key == "Megapixel" and value:
                            overlay_text_content += f"Megapixel: {value}\n"

                overlay_text(filepath, overlay_text_content, (10, 10), OUTPUT_DIRECTORY)

        # Create directory metadata:
        working_directory_name = os.path.basename(INPUT_DIRECTORY)
        directory_metadata_path = os.path.join(
            OUTPUT_DIRECTORY, working_directory_name + "_MD." + METADATA_FORMAT
        )
        if CREATE_METADATA_FILE:
            write_metadata(directory_metadata, directory_metadata_path)


if __name__ == "__main__":
    main()
